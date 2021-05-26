import os
from openpyxl import load_workbook
from openpyxl import Workbook

excel_directory_move = os.getcwd()
print(excel_directory_move)
wb1 = load_workbook('{}\\E0000237-Auxiliary_Module_1-04_03_21.xlsx'.format(excel_directory_move))
row_start = 13
row_end = 15
col = 3

class ExcelManagement():

    def __init__(self, wb):
        self.ws = wb.active
        
    def get_value(self, row_cursor, col_cursor):
        temp_location = self.ws.cell(row=row_cursor, column=col_cursor)
        temp_cell_value= temp_location.value
        return (temp_cell_value)

    def list_creator(self, row_start, row_end, col):
        wb_list = []
        for value in range (row_start, row_end):
            wb_list.append(self.get_value(value, col))
        return wb_list

    def main_program(self, row_start, row_end, column_list):
        for 

class ListManagement():

    def __init__(self):
        self.source_listA = ExcelManagement(wb1).list_creator(row_start,row_end,col)
        self.source_listB = ExcelManagement(wb1).list_creator(row_start,row_end,col)

    def column_list_creation(colA, colB, *args):
        

    def list_comparator_rank1(self, self.source_listA, self.source_listB):
        destination_list = []
        temp_list = []
        for value_listA in self.source_listA:
            for value_listB in self.source_listB:
                if value_listA == value_listB:
                    temp_list.append(self.get_value(value, col))
        return wb_list

print(ExcelManagement(wb1).list_creator(13,15,3))
